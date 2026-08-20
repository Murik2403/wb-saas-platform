from __future__ import annotations

from io import BytesIO
import re
from typing import Any

import pandas as pd


def _numeric(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series(0.0, index=frame.index, dtype=float)
    return pd.to_numeric(frame[column], errors="coerce").fillna(0.0)


def _find_column(columns: list[str], startswith: str) -> str | None:
    needle = startswith.casefold()
    for col in columns:
        if str(col).strip().casefold().startswith(needle):
            return str(col)
    return None


def parse_income_expense_report(source: Any) -> tuple[dict[str, Any], pd.DataFrame]:
    """Parse the WB report «Доходы и расходы» exported as XLSX.

    Returns aggregate reference metrics and the normalized product-level table.
    The function is deliberately tolerant to missing optional columns because WB
    periodically changes the report layout.
    """
    if hasattr(source, "seek"):
        source.seek(0)
    excel = pd.ExcelFile(source, engine="openpyxl")
    sheet_names = excel.sheet_names
    detail_sheet = "Детальная информация" if "Детальная информация" in sheet_names else sheet_names[-1]
    detail = pd.read_excel(excel, sheet_name=detail_sheet, header=1)
    detail.columns = [str(c).strip() for c in detail.columns]
    detail = detail.dropna(how="all")

    period_start = None
    period_end = None
    if "Общая информация" in sheet_names:
        general = pd.read_excel(excel, sheet_name="Общая информация", header=None)
        text = " ".join(str(v) for v in general.to_numpy().ravel() if pd.notna(v))
        match = re.search(r"(20\d{2}-\d{2}-\d{2}).*?(20\d{2}-\d{2}-\d{2})", text)
        if match:
            period_start = pd.to_datetime(match.group(1)).date()
            period_end = pd.to_datetime(match.group(2)).date()

    columns = list(detail.columns)
    sales_col = _find_column(columns, "Продажи, ₽")
    returns_col = _find_column(columns, "Возвраты, ₽")
    sale_units_col = _find_column(columns, "Продажи, шт")
    return_units_col = _find_column(columns, "Возвраты, шт")
    logistics_col = _find_column(columns, "Логистика, ₽")
    penalties_col = _find_column(columns, "Штрафы, ₽")
    commission_col = _find_column(columns, "Комиссия WB, ₽")
    acquiring_col = _find_column(columns, "Эквайринг, ₽")
    losses_col = _find_column(columns, "Потери, подмены")
    surcharges_col = _find_column(columns, "Доплаты, ₽")
    loyalty_col = _find_column(columns, "Программа лояльности, ₽")
    total_col = _find_column(columns, "Итог, ₽")

    sales = _numeric(detail, sales_col or "")
    returns = _numeric(detail, returns_col or "")
    sale_units = _numeric(detail, sale_units_col or "")
    return_units = _numeric(detail, return_units_col or "")
    logistics = _numeric(detail, logistics_col or "")
    penalties = _numeric(detail, penalties_col or "")
    commission = _numeric(detail, commission_col or "")
    acquiring = _numeric(detail, acquiring_col or "")
    losses = _numeric(detail, losses_col or "")
    surcharges = _numeric(detail, surcharges_col or "")
    loyalty = _numeric(detail, loyalty_col or "")
    total = _numeric(detail, total_col or "")

    metrics: dict[str, Any] = {
        "period_start": period_start,
        "period_end": period_end,
        "rows": int(len(detail)),
        "sales_amount": float(sales.sum()),
        "returns_amount": float(abs(returns.sum())),
        "gross_sales": float((sales + returns).sum()),
        "sale_units": float(sale_units.sum()),
        "return_units": float(return_units.sum()),
        "net_units": float((sale_units - return_units).sum()),
        "logistics": float(abs(logistics.sum())),
        "penalties": float(abs(penalties.sum())),
        "commission": float(abs(commission.sum())),
        "acquiring": float(abs(acquiring.sum())),
        "losses": float(abs(losses.sum())),
        "surcharges": float(surcharges.sum()),
        "loyalty": float(loyalty.sum()),
        "product_result": float(total.sum()),
    }

    # A compact, normalized table is useful for an Excel export and manual audit.
    keep = [
        "Артикул продавца", "Артикул WB", "Название", "Предмет", "Бренд",
        total_col, sales_col, sale_units_col, returns_col, return_units_col,
        logistics_col, penalties_col, commission_col, acquiring_col,
        losses_col, surcharges_col, loyalty_col,
    ]
    keep = [c for c in keep if c and c in detail.columns]
    normalized = detail[keep].copy()
    return metrics, normalized


def build_reconciliation(financial: dict[str, float], wb: dict[str, Any]) -> pd.DataFrame:
    rows = [
        ("Выкупы нетто, ₽", financial.get("gross_sales", 0.0), wb.get("gross_sales", 0.0), "money"),
        ("Продажи, шт", financial.get("sale_units", 0.0), wb.get("sale_units", 0.0), "units"),
        ("Возвраты, шт", financial.get("return_units", 0.0), wb.get("return_units", 0.0), "units"),
        ("Продано нетто, шт", financial.get("net_units", 0.0), wb.get("net_units", 0.0), "units"),
        ("Логистика, ₽", financial.get("logistics", 0.0), wb.get("logistics", 0.0), "money"),
        ("Штрафы, ₽", financial.get("penalties", 0.0), wb.get("penalties", 0.0), "money"),
        ("Эквайринг, ₽", financial.get("acquiring", 0.0), wb.get("acquiring", 0.0), "money"),
    ]
    output = []
    for article, dashboard_value, wb_value, kind in rows:
        dashboard_value = float(dashboard_value or 0)
        wb_value = float(wb_value or 0)
        delta = dashboard_value - wb_value
        if kind == "units":
            tolerance = 0.5
        else:
            tolerance = max(100.0, abs(wb_value) * 0.001)
        status = "Совпадает" if abs(delta) <= tolerance else "Есть отклонение"
        output.append({
            "Показатель": article,
            "Дашборд": dashboard_value,
            "Отчёт WB": wb_value,
            "Отклонение": delta,
            "Отклонение, %": (delta / wb_value * 100) if wb_value else 0.0,
            "Статус": status,
        })
    return pd.DataFrame(output)


def build_finance_excel(
    *,
    start: Any,
    end: Any,
    financial: dict[str, float],
    ad_spend: float,
    financial_products: pd.DataFrame,
    expense_rows: pd.DataFrame,
    reconciliation: pd.DataFrame | None = None,
    wb_detail: pd.DataFrame | None = None,
    store_operations: pd.DataFrame | None = None,
) -> bytes:
    """Build a formatted finance/unit-economics workbook for download."""
    buffer = BytesIO()
    summary = pd.DataFrame([
        ["Период", f"{start:%d.%m.%Y}–{end:%d.%m.%Y}"],
        ["Выкупы по цене покупателя, ₽", financial.get("gross_sales", 0.0)],
        ["К перечислению за товар, ₽", financial.get("to_pay", 0.0)],
        ["Расходы WB, ₽", financial.get("wb_expenses", 0.0)],
        ["Реклама, ₽", ad_spend],
        ["Себестоимость, ₽", financial.get("cost", 0.0)],
        ["Прибыль до налога, ₽", financial.get("profit", 0.0)],
        ["Продано нетто, шт", financial.get("net_units", 0.0)],
        ["Покрытие себестоимости, %", financial.get("cost_coverage", 0.0)],
        ["Распределено общих расходов, ₽", financial.get("common_expense_pool", 0.0)],
        ["Удержания на уровне магазина, ₽", financial.get("unallocated_deductions", 0.0)],
        ["Остаток прибыли на уровне магазина, ₽", financial.get("remaining_store_impact", 0.0)],
    ], columns=["Показатель", "Значение"])

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Сводка", index=False, startrow=2)
        expense_rows.to_excel(writer, sheet_name="Статьи расходов", index=False)
        financial_products.to_excel(writer, sheet_name="Юнит-экономика", index=False)
        if reconciliation is not None and not reconciliation.empty:
            reconciliation.to_excel(writer, sheet_name="Сверка WB", index=False)
        if wb_detail is not None and not wb_detail.empty:
            wb_detail.to_excel(writer, sheet_name="Выгрузка WB", index=False)
        if store_operations is not None and not store_operations.empty:
            store_operations.to_excel(writer, sheet_name="Операции магазина", index=False)

        methodology = pd.DataFrame([
            ["Общая прибыль магазина", "К перечислению + доплаты WB − расходы WB − реклама − себестоимость"],
            ["Маржинальная прибыль товара", "Доход товара после прямых расходов WB, рекламы и себестоимости"],
            ["Расчётная прибыль товара", "Маржинальная прибыль − доля общих распределяемых расходов"],
            ["Распределение общих расходов", "По доле выручки активного товара; при отсутствии выручки — по количеству"],
            ["Не распределяется", "Удержания магазина и технические/неактивные карточки остаются отдельными строками"],
            ["Расходы WB", "Логистика + дополнительная логистика + хранение + штрафы + удержания + приёмка"],
            ["Рентабельность затрат, %", "Расчётная прибыль / (прямые расходы WB + реклама + себестоимость + распределённые расходы)"],
            ["Примечание", "Налог, аренда, зарплата, электричество и амортизация оборудования в расчёт не включены."],
        ], columns=["Показатель", "Методика"])
        methodology.to_excel(writer, sheet_name="Методика", index=False)

        wb = writer.book
        dark = "111827"
        purple = "7C3AED"
        light = "EDE9FE"
        green = "DCFCE7"
        red = "FEE2E2"
        white = "FFFFFF"
        thin = "D1D5DB"

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.formatting.rule import CellIsRule

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=dark)
                cell.font = Font(color=white, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border(bottom=Side(style="hair", color=thin))
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col_cells in ws.columns:
                values = [str(c.value) for c in col_cells if c.value is not None]
                width = min(max((len(v) for v in values), default=10) + 2, 38)
                ws.column_dimensions[col_cells[0].column_letter].width = width

        ws = wb["Сводка"]
        ws["A1"] = "MARKETSHELPER — финансовая сводка"
        ws.merge_cells("A1:B1")
        ws["A1"].fill = PatternFill("solid", fgColor=purple)
        ws["A1"].font = Font(color=white, bold=True, size=15)
        ws["A1"].alignment = Alignment(horizontal="center")
        for row in range(4, ws.max_row + 1):
            label = str(ws.cell(row, 1).value or "")
            value_cell = ws.cell(row, 2)
            if label == "Период":
                value_cell.number_format = "General"
            elif label == "Продано нетто, шт":
                value_cell.number_format = '#,##0'
            elif label == "Покрытие себестоимости, %":
                value_cell.number_format = '0.0%'
                value_cell.value = float(financial.get("cost_coverage", 0.0)) / 100.0
            else:
                value_cell.number_format = '#,##0.00 [$₽-ru-RU]'

        money_headers = {
            "Выкупы по цене покупателя", "Продажи по отчёту", "К перечислению за товар",
            "Прямые расходы WB", "Реклама", "Себестоимость ед.", "Себестоимость",
            "Маржинальная прибыль", "Маржинальная прибыль на ед.",
            "Распределено общих расходов", "Расчётная прибыль", "Расчётная прибыль на ед."
        }
        unit_ws = wb["Юнит-экономика"]
        header_map = {cell.value: cell.column for cell in unit_ws[1]}
        for header, col_idx in header_map.items():
            if header in money_headers:
                for row in range(2, unit_ws.max_row + 1):
                    unit_ws.cell(row, col_idx).number_format = '#,##0.00 [$₽-ru-RU]'
            elif header and ("%" in str(header) or header in {"Маржа, %", "Рентабельность затрат, %", "Доля прибыли, %"}):
                for row in range(2, unit_ws.max_row + 1):
                    unit_ws.cell(row, col_idx).number_format = '0.0%'
                    value = unit_ws.cell(row, col_idx).value
                    if isinstance(value, (int, float)):
                        unit_ws.cell(row, col_idx).value = value / 100.0
        profit_col = header_map.get("Расчётная прибыль")
        if profit_col:
            letter = unit_ws.cell(1, profit_col).column_letter
            rng = f"{letter}2:{letter}{unit_ws.max_row}"
            unit_ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=red)))
            unit_ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor=green)))

        if "Сверка WB" in wb.sheetnames:
            recon_ws = wb["Сверка WB"]
            status_col = next((c.column for c in recon_ws[1] if c.value == "Статус"), None)
            if status_col:
                for row in range(2, recon_ws.max_row + 1):
                    cell = recon_ws.cell(row, status_col)
                    cell.fill = PatternFill("solid", fgColor=green if cell.value == "Совпадает" else red)

        for ws_name in ["Статьи расходов", "Сверка WB", "Операции магазина"]:
            if ws_name in wb.sheetnames:
                wsx = wb[ws_name]
                for row in wsx.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00 [$₽-ru-RU]'

    return buffer.getvalue()
